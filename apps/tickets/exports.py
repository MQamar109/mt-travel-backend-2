from decimal import Decimal
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from apps.core.export_pdf import pdf_cell_paragraph, pdf_row_cells, proportional_col_widths, text_column_indices
from apps.core.running_balance import apply_balance_step
from .utils import get_display_amount

HEADER_COLOR_HEX = '1F4E79'
ALT_ROW_HEX = 'D9E1F2'


def _build_rows(tickets, display_currency, show_vendor, show_dep_date, show_tickets_count):
    """Returns (headers, data_rows) with running balance applied."""
    headers = ['No', 'Issued Date', 'Invoice No']
    if show_vendor:
        headers.append('Vendor')
    headers.append('Customer Name')
    if show_dep_date:
        headers.append('Dep Date')
    if show_tickets_count:
        headers.append('No of Tickets')
    headers += [f'Amount ({display_currency})', 'Payment Type', f'Balance ({display_currency})']

    data_rows = []
    running_balance = None

    for i, ticket in enumerate(tickets, 1):
        amount = get_display_amount(
            ticket.total_amount, ticket.currency, ticket.exchange_rate, display_currency
        )

        running_balance = apply_balance_step(running_balance, ticket.payment_type, amount)

        row = [i, str(ticket.issued_date), ticket.invoice_no]
        if show_vendor:
            row.append(ticket.vendor.name)
        row.append(ticket.customer_name)
        if show_dep_date:
            row.append(str(ticket.departure_date))
        if show_tickets_count:
            row.append(ticket.no_of_tickets)
        row += [float(amount), ticket.payment_type, float(running_balance)]
        data_rows.append(row)

    return headers, data_rows


def generate_excel(tickets, display_currency, meta, options=None):
    options = options or {}
    show_vendor = options.get('show_vendor', True)
    show_dep_date = options.get('show_dep_date', True)
    show_tickets_count = options.get('show_tickets_count', True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tickets'

    # Title row
    ws.merge_cells('A1:J1')
    ws['A1'].value = 'Ticket Report'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Meta row (duration + optional vendor)
    meta_parts = [meta.get('duration', 'All Dates')]
    if meta.get('vendor_name'):
        meta_parts.append(f"Vendor: {meta['vendor_name']}")
    meta_parts.append(f'Currency: {display_currency}')
    ws.merge_cells('A2:J2')
    ws['A2'].value = '   |   '.join(meta_parts)
    ws['A2'].font = Font(size=11, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.append([])  # blank spacer row

    headers, data_rows = _build_rows(
        tickets, display_currency, show_vendor, show_dep_date, show_tickets_count
    )

    # Header row
    header_row_idx = ws.max_row + 1
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=HEADER_COLOR_HEX, end_color=HEADER_COLOR_HEX, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Data rows
    for i, row in enumerate(data_rows):
        ws.append(row)
        row_idx = ws.max_row
        fill_color = ALT_ROW_HEX if i % 2 == 0 else 'FFFFFF'
        for col_idx in range(1, len(row) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            header = headers[col_idx - 1].lower() if col_idx <= len(headers) else ''
            wrap = 'name' in header or 'customer' in header or 'vendor' in header
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=wrap,
            )
            if isinstance(cell.value, float):
                cell.number_format = '#,##0.00'

    # Column widths — cap wide text columns, ensure amount/type/balance visible
    for col_idx, header in enumerate(headers, 1):
        letter = get_column_letter(col_idx)
        h = header.lower()
        if 'name' in h or 'customer' in h:
            ws.column_dimensions[letter].width = 28
        elif 'amount' in h or 'balance' in h:
            ws.column_dimensions[letter].width = 14
        elif 'payment' in h or 'type' in h:
            ws.column_dimensions[letter].width = 12
        else:
            max_len = max(
                (len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(header_row_idx, ws.max_row + 1)),
                default=10,
            )
            ws.column_dimensions[letter].width = min(max_len + 3, 22)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_pdf(tickets, display_currency, meta, options=None):
    options = options or {}
    show_vendor = options.get('show_vendor', True)
    show_dep_date = options.get('show_dep_date', True)
    show_tickets_count = options.get('show_tickets_count', True)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        rightMargin=1*cm, leftMargin=1*cm,
        topMargin=1.5*cm, bottomMargin=1*cm,
    )

    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle('ReportTitle', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16, spaceAfter=4)
    elements.append(Paragraph('Ticket Report', title_style))

    # Meta
    meta_parts = [meta.get('duration', 'All Dates')]
    if meta.get('vendor_name'):
        meta_parts.append(f"Vendor: {meta['vendor_name']}")
    meta_parts.append(f'Currency: {display_currency}')
    meta_style = ParagraphStyle('Meta', parent=styles['Normal'], alignment=TA_CENTER, fontSize=10)
    elements.append(Paragraph('   |   '.join(meta_parts), meta_style))
    elements.append(Spacer(1, 14))

    headers, data_rows = _build_rows(
        tickets, display_currency, show_vendor, show_dep_date, show_tickets_count
    )

    text_indices = text_column_indices(headers)
    pdf_rows = [pdf_row_cells(row, text_indices) for row in data_rows]

    table_data = [headers] + pdf_rows

    header_bg = colors.HexColor(f'#{HEADER_COLOR_HEX}')
    alt_bg = colors.HexColor(f'#{ALT_ROW_HEX}')

    row_backgrounds = []
    for i in range(1, len(pdf_rows) + 1):
        bg = colors.white if i % 2 != 0 else alt_bg
        row_backgrounds.append(('BACKGROUND', (0, i), (-1, i), bg))

    usable_width = landscape(A4)[0] - 2 * cm
    col_widths = proportional_col_widths(headers, usable_width)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), header_bg),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ] + row_backgrounds)

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(style)
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    return buffer
