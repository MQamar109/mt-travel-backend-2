from decimal import Decimal
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from apps.core.running_balance import apply_balance_step
from apps.tickets.utils import get_display_amount

# ── Palette ──────────────────────────────────────────────────────────────────
H1_HEX = '1F3A52'   # darkest  – top group row
H2_HEX = '25587D'   # medium   – sub-group row
H3_HEX = '2E75B6'   # lighter  – unit label row
ALT_HEX = 'D9E1F2'  # alternating data row tint

H1_C = colors.HexColor(f'#{H1_HEX}')
H2_C = colors.HexColor(f'#{H2_HEX}')
H3_C = colors.HexColor(f'#{H3_HEX}')
ALT_C = colors.HexColor(f'#{ALT_HEX}')

# 24 columns: A‥X
# A=#  B=Vendor  C=In  D=Out  E=Hotel  F=ResNo  G=Guest  H=Nts
# I=SGL qty  J=DBL qty  K=TRL qty  L=QUAD qty
# M=SGL rate N=DBL rate O=TRL rate P=QUAD rate
# Q=BF rate  R=LU rate  S=DI rate
# T=Room Total  U=Meal Total  V=Grand Total  W=Type  X=Balance
NUM_COLS = 24


def _cv(value, record_currency, exchange_rate, display_currency):
    return float(get_display_amount(value, record_currency, exchange_rate, display_currency))


def _build_data_rows(hotels, display_currency):
    rows = []
    running_balance = None

    for i, h in enumerate(hotels, 1):
        c = lambda val, _h=h: _cv(val, _h.currency, _h.exchange_rate, display_currency)
        grand_total = c(h.total_amount)

        running_balance = apply_balance_step(running_balance, h.payment_type, grand_total)

        rows.append([
            i,
            h.vendor.name,
            str(h.check_in),
            str(h.check_out),
            h.hotel_name,
            h.reservation_no,
            h.guest_name,
            h.nights,
            # Room qty — None when 0 (renders blank / "—")
            h.single_qty or None,
            h.double_qty or None,
            h.triple_qty or None,
            h.quad_qty or None,
            # Room rates — only meaningful when the room type is used
            c(h.single_rate) if h.single_qty else None,
            c(h.double_rate) if h.double_qty else None,
            c(h.triple_rate) if h.triple_qty else None,
            c(h.quad_rate) if h.quad_qty else None,
            # Meal rates — only when meals are enabled
            c(h.bf_rate or 0) if h.meals else None,
            c(h.lu_rate or 0) if h.meals else None,
            c(h.di_rate or 0) if h.meals else None,
            # Totals
            c(h.total_room_amount),
            c(h.total_meal_amount) if h.meals else None,
            grand_total,
            h.payment_type,
            float(running_balance),
        ])

    return rows


# ── Excel ────────────────────────────────────────────────────────────────────

def _hcell(ws, addr, value, bg_hex):
    cell = ws[addr]
    cell.value = value
    cell.font = Font(bold=True, color='FFFFFF', size=8)
    cell.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def generate_excel(hotels, display_currency, meta):
    data_rows = _build_data_rows(hotels, display_currency)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Hotels'

    last_col = get_column_letter(NUM_COLS)  # 'X'

    # Title
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'].value = 'Hotel Report'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 22

    # Meta (duration / vendor / currency)
    meta_parts = [meta.get('duration', 'All Dates')]
    if meta.get('vendor_name'):
        meta_parts.append(f"Vendor: {meta['vendor_name']}")
    meta_parts.append(f"Currency: {display_currency}")
    ws.merge_cells(f'A2:{last_col}2')
    ws['A2'].value = '   |   '.join(meta_parts)
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.append([])  # Row 3: spacer

    H1, H2, H3 = 4, 5, 6

    # Tall columns: set value+style FIRST, then merge rows H1–H3.
    # openpyxl preserves cell content when the value is written before merge_cells.
    TALL_COLS = [
        ('A', '#'), ('B', 'VENDOR'), ('E', 'HOTEL'), ('F', 'RES. NO'),
        ('G', 'GUEST'), ('H', 'NTS'),
        ('T', 'ROOM TOTAL'), ('U', 'MEAL TOTAL'), ('V', 'GRAND TOTAL'),
        ('W', 'TYPE'), ('X', 'BALANCE'),
    ]
    for col, label in TALL_COLS:
        cell = ws[f'{col}{H1}']
        cell.value = label
        cell.font = Font(bold=True, color='FFFFFF', size=8)
        cell.fill = PatternFill(start_color=H1_HEX, end_color=H1_HEX, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'{col}{H1}:{col}{H3}')

    # ── Row H1: group headers (horizontal merges only) ──
    ws.merge_cells(f'C{H1}:D{H1}')
    _hcell(ws, f'C{H1}', 'CHECKED DATE', H1_HEX)

    ws.merge_cells(f'I{H1}:L{H1}')
    _hcell(ws, f'I{H1}', 'NO. OF ROOMS', H1_HEX)

    ws.merge_cells(f'M{H1}:P{H1}')
    _hcell(ws, f'M{H1}', 'ROOM RATE / NIGHT', H1_HEX)

    ws.merge_cells(f'Q{H1}:S{H1}')
    _hcell(ws, f'Q{H1}', 'MEAL RATE / GUEST', H1_HEX)

    # ── Row H2: sub-group headers ──
    for col, label in [
        ('C', 'IN'), ('D', 'OUT'),
        ('I', 'SGL'), ('J', 'DBL'), ('K', 'TRL'), ('L', 'QUAD'),
        ('M', 'SGL'), ('N', 'DBL'), ('O', 'TRL'), ('P', 'QUAD'),
        ('Q', 'BF'),  ('R', 'LU'),  ('S', 'DI'),
    ]:
        _hcell(ws, f'{col}{H2}', label, H2_HEX)

    # ── Row H3: unit labels ──
    for col, label in [
        ('C', 'DATE'), ('D', 'DATE'),
        ('I', 'QTY'),    ('J', 'QTY'),    ('K', 'QTY'),    ('L', 'QTY'),
        ('M', '/NIGHT'), ('N', '/NIGHT'), ('O', '/NIGHT'), ('P', '/NIGHT'),
        ('Q', '/NIGHT'), ('R', '/NIGHT'), ('S', '/NIGHT'),
    ]:
        _hcell(ws, f'{col}{H3}', label, H3_HEX)

    for r in [H1, H2, H3]:
        ws.row_dimensions[r].height = 18

    # ── Data rows ──
    for i, row in enumerate(data_rows):
        ws.append(row)
        row_idx = ws.max_row
        fill_hex = ALT_HEX if i % 2 == 0 else 'FFFFFF'
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=8)
            if isinstance(val, float):
                cell.number_format = '#,##0.00'
            elif isinstance(val, int) and col_idx > 1:
                cell.number_format = '#,##0'

    # ── Column widths ──
    for col_letter, width in {
        'A': 5,  'B': 16, 'C': 12, 'D': 12, 'E': 20, 'F': 10,
        'G': 14, 'H': 5,
        'I': 6,  'J': 6,  'K': 6,  'L': 6,
        'M': 11, 'N': 11, 'O': 11, 'P': 11,
        'Q': 9,  'R': 9,  'S': 9,
        'T': 13, 'U': 11, 'V': 13, 'W': 8,  'X': 13,
    }.items():
        ws.column_dimensions[col_letter].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ── PDF ──────────────────────────────────────────────────────────────────────

# Column widths in points for landscape A4 (~825pt usable)
_PDF_COL_WIDTHS = [
    18,           # #
    55,           # Vendor
    38, 38,       # In / Out
    60,           # Hotel
    38,           # Res No
    52,           # Guest
    16,           # Nts
    18, 18, 18, 18,        # Room qty
    36, 36, 36, 36,        # Room rates
    28, 28, 28,            # Meal rates
    44, 36, 44,            # Room Total / Meal Total / Grand Total
    30,           # Type
    44,           # Balance
]  # total ≈ 813pt


def generate_pdf(hotels, display_currency, meta):
    data_rows = _build_data_rows(hotels, display_currency)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        rightMargin=0.3 * cm, leftMargin=0.3 * cm,
        topMargin=1.5 * cm, bottomMargin=1 * cm,
    )

    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle(
        'HotelTitle', parent=styles['Heading1'],
        alignment=TA_CENTER, fontSize=14, spaceAfter=4,
    )
    elements.append(Paragraph('Hotel Report', title_style))

    meta_parts = [meta.get('duration', 'All Dates')]
    if meta.get('vendor_name'):
        meta_parts.append(f"Vendor: {meta['vendor_name']}")
    meta_parts.append(f"Currency: {display_currency}")
    elements.append(Paragraph(
        '   |   '.join(meta_parts),
        ParagraphStyle('HotelMeta', parent=styles['Normal'], alignment=TA_CENTER, fontSize=9),
    ))
    elements.append(Spacer(1, 10))

    # 3-row header with vertical SPANs for tall columns.
    # Key: apply H1_C to the ENTIRE header block first, then override only the
    # sub-header cells with H2/H3. Span cells inherit H1_C from that wide rule.
    #
    # Tall col indices (0-based):
    #   0=#  1=VENDOR  4=HOTEL  5=RES.NO  6=GUEST  7=NTS
    #   19=ROOM TOTAL  20=MEAL TOTAL  21=GRAND TOTAL  22=TYPE  23=BALANCE
    h_row1 = [
        '#', 'VENDOR', 'CHECKED DATE', '', 'HOTEL', 'RES. NO', 'GUEST', 'NTS',
        'NO. OF ROOMS', '', '', '',
        'ROOM RATE / NIGHT', '', '', '',
        'MEAL RATE / GUEST', '', '',
        'ROOM TOTAL', 'MEAL TOTAL', 'GRAND TOTAL', 'TYPE', 'BALANCE',
    ]
    h_row2 = [
        '', '', 'IN', 'OUT', '', '', '', '',
        'SGL', 'DBL', 'TRL', 'QUAD',
        'SGL', 'DBL', 'TRL', 'QUAD',
        'BF',  'LU',  'DI',
        '', '', '', '', '',
    ]
    h_row3 = [
        '', '', 'DATE', 'DATE', '', '', '', '',
        'QTY', 'QTY', 'QTY', 'QTY',
        '/NIGHT', '/NIGHT', '/NIGHT', '/NIGHT',
        '/NIGHT', '/NIGHT', '/NIGHT',
        '', '', '', '', '',
    ]

    pdf_data = []
    for row in data_rows:
        pdf_data.append([
            '—' if v is None
            else (f'{v:,.2f}' if isinstance(v, float) else str(v))
            for v in row
        ])

    table_data = [h_row1, h_row2, h_row3] + pdf_data

    # Vertical SPANs for tall columns + horizontal SPANs for group headers
    spans = [
        # Tall columns: merge all 3 header rows → text appears vertically centred
        ('SPAN', (0, 0),  (0, 2)),   # #
        ('SPAN', (1, 0),  (1, 2)),   # VENDOR
        ('SPAN', (4, 0),  (4, 2)),   # HOTEL
        ('SPAN', (5, 0),  (5, 2)),   # RES. NO
        ('SPAN', (6, 0),  (6, 2)),   # GUEST
        ('SPAN', (7, 0),  (7, 2)),   # NTS
        ('SPAN', (19, 0), (19, 2)),  # ROOM TOTAL
        ('SPAN', (20, 0), (20, 2)),  # MEAL TOTAL
        ('SPAN', (21, 0), (21, 2)),  # GRAND TOTAL
        ('SPAN', (22, 0), (22, 2)),  # TYPE
        ('SPAN', (23, 0), (23, 2)),  # BALANCE
        # Horizontal group headers (row 0)
        ('SPAN', (2, 0),  (3, 0)),   # CHECKED DATE
        ('SPAN', (8, 0),  (11, 0)),  # NO. OF ROOMS
        ('SPAN', (12, 0), (15, 0)),  # ROOM RATE / NIGHT
        ('SPAN', (16, 0), (18, 0)),  # MEAL RATE / GUEST
    ]

    # Apply H1 to the ENTIRE header block first, then override non-span cells.
    # Span cells inherit H1_C from this wide rule; individual cell overrides on
    # span members (other than top-left) would conflict — avoid them entirely.
    header_bg = [
        ('BACKGROUND', (0, 0),  (-1, 2),  H1_C),  # all 3 header rows → H1 dark
        # Sub-header row (row 1) — non-span cells only
        ('BACKGROUND', (2, 1),  (3, 1),  H2_C),   # IN, OUT
        ('BACKGROUND', (8, 1),  (11, 1), H2_C),   # SGL–QUAD qty
        ('BACKGROUND', (12, 1), (15, 1), H2_C),   # SGL–QUAD rate
        ('BACKGROUND', (16, 1), (18, 1), H2_C),   # BF, LU, DI
        # Unit label row (row 2) — non-span cells only
        ('BACKGROUND', (2, 2),  (3, 2),  H3_C),   # DATE, DATE
        ('BACKGROUND', (8, 2),  (11, 2), H3_C),   # QTY×4
        ('BACKGROUND', (12, 2), (15, 2), H3_C),   # /NIGHT×4
        ('BACKGROUND', (16, 2), (18, 2), H3_C),   # /NIGHT×3
    ]

    alt_rows = [
        ('BACKGROUND', (0, i + 3), (-1, i + 3), colors.white if i % 2 != 0 else ALT_C)
        for i in range(len(pdf_data))
    ]

    t = Table(table_data, colWidths=_PDF_COL_WIDTHS, repeatRows=3)
    t.setStyle(TableStyle(spans + header_bg + [
        ('TEXTCOLOR',     (0, 0), (-1, 2), colors.white),
        ('FONTNAME',      (0, 0), (-1, 2), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 2), 6),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE',      (0, 3), (-1, -1), 6),
        ('GRID',          (0, 0), (-1, -1), 0.3, colors.grey),
        ('TOPPADDING',    (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING',   (0, 0), (-1, -1), 2),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 2),
    ] + alt_rows))

    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    return buffer
